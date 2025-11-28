"""Read WAYBILL RECORD.xlsx and print waybill numbers where column C is empty.

Usage:
    python Database\read_waybills.py

Functions:
    get_next_unclaimed_waybill() -> str or None
    mark_waybill_used(waybill_no: str) -> bool
    save_waybills_to_files() -> None

If `openpyxl` is not installed, the script will print pip install instructions.
"""
from pathlib import Path
import sys
import json
import csv
from datetime import datetime
import time
from typing import Optional
import os

EXCEL_PATH = Path(__file__).parent / "WAYBILL RECORD.xlsx"
JSON_PATH = Path(__file__).parent / "waybills.json"
CSV_PATH = Path(__file__).parent / "waybills.csv"
START_ROW = 7  # Rows 1-6 are headers/title


def find_empty_column_c_waybills(path: Path):
    """Find all waybills where column C is empty, starting from START_ROW."""
    try:
        import openpyxl
    except Exception:
        print("Missing dependency: openpyxl is required to read .xlsx files.")
        print("Install it with: python -m pip install openpyxl")
        return None

    if not path.exists():
        print(f"Excel file not found: {path}")
        return None

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    waybills = []
    # Iterate rows starting from START_ROW
    for row in range(START_ROW, ws.max_row + 1):
        # Column C is column index 3
        cell_c = ws.cell(row=row, column=3).value
        # treat empty, None, or whitespace-only as empty
        if cell_c is None or (isinstance(cell_c, str) and cell_c.strip() == ""):
            # Read waybill from column A (merged A:B typically stores it)
            wb_cell = ws.cell(row=row, column=1).value
            if wb_cell is None:
                continue
            # Normalize to string and strip
            wb_str = str(wb_cell).strip()
            if wb_str and wb_str.isdigit():  # Only numeric waybills
                waybills.append({"waybill_no": wb_str, "row": row, "timestamp": None})

    return waybills


def save_waybills_to_files():
    """Find unclaimed waybills and save to JSON and CSV."""
    waybills = find_empty_column_c_waybills(EXCEL_PATH)
    if waybills is None:
        return False

    # Save to JSON
    with open(JSON_PATH, "w") as f:
        json.dump(waybills, f, indent=2)
    print(f"Saved {len(waybills)} waybills to {JSON_PATH}")

    # Save to CSV
    with open(CSV_PATH, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=["waybill_no", "row", "timestamp"])
        writer.writeheader()
        writer.writerows(waybills)
    print(f"Saved {len(waybills)} waybills to {CSV_PATH}")

    return True


def get_next_unclaimed_waybill() -> Optional[str]:
    """Return the first unclaimed waybill number from JSON, or None if all are claimed."""
    if not JSON_PATH.exists():
        # Generate JSON if it doesn't exist
        save_waybills_to_files()

    try:
        with open(JSON_PATH, "r") as f:
            waybills = json.load(f)
    except Exception as e:
        print(f"Error reading {JSON_PATH}: {e}")
        return None

    # Find first one with no timestamp (unclaimed)
    for wb in waybills:
        if wb.get("timestamp") is None:
            return wb.get("waybill_no")

    return None  # All claimed


def mark_waybill_used(waybill_no: str) -> bool:
    """Mark a waybill as used by adding a timestamp to JSON and writing to Excel column C."""
    if not JSON_PATH.exists():
        save_waybills_to_files()

    # Update JSON
    try:
        with open(JSON_PATH, "r") as f:
            waybills = json.load(f)
    except Exception as e:
        print(f"Error reading {JSON_PATH}: {e}")
        return False

    found = False
    timestamp = datetime.now().isoformat()
    for wb in waybills:
        if wb.get("waybill_no") == waybill_no:
            wb["timestamp"] = timestamp
            found = True
            break

    if not found:
        print(f"Waybill {waybill_no} not found in {JSON_PATH}")
        return False

    # Write back to JSON
    try:
        with open(JSON_PATH, "w") as f:
            json.dump(waybills, f, indent=2)
    except Exception as e:
        print(f"Error writing to {JSON_PATH}: {e}")
        return False

    # Write to Excel column C
    try:
        import openpyxl

        # Try several times to write the Excel file in case it's briefly locked by Excel.
        max_attempts = 5
        attempt = 0
        last_exc = None
        def _is_locked(path: Path) -> bool:
            # Try platform-specific non-blocking lock (Windows: msvcrt)
            try:
                import msvcrt
                with open(path, "r+b") as fh:
                    try:
                        msvcrt.locking(fh.fileno(), msvcrt.LK_NBLCK, 1)
                        msvcrt.locking(fh.fileno(), msvcrt.LK_UNLCK, 1)
                        return False
                    except OSError:
                        return True
            except Exception:
                # Fallback: try append open (may not always detect locks)
                try:
                    f = open(path, "a")
                    f.close()
                    return False
                except Exception:
                    return True

        # If the Excel file is currently open/locked, skip the Excel write to avoid hanging.
        try:
            f = open(EXCEL_PATH, "a")
            f.close()
        except Exception:
            print(f"Warning: Excel file appears locked. Skipping Excel write for {waybill_no}.")
            # CSV/JSON already updated above; return success so the workflow continues.
            return True

        try:
            wb_file = openpyxl.load_workbook(EXCEL_PATH)
            ws = wb_file.active
            for wb in waybills:
                if wb.get("waybill_no") == waybill_no:
                    row = wb.get("row")
                    ws.cell(row=row, column=3).value = "TRANSFER"
                    break
            wb_file.save(EXCEL_PATH)
            print(f"Marked waybill {waybill_no} as used in Excel (row {row}, column C set to TRANSFER)")
        except Exception as e:
            print(f"Error writing to Excel: {e}")
            print("Excel write skipped; please close the file and re-run if you need Column C updated.")
            return True
        if last_exc:
            print(f"Warning: Could not write to Excel after {max_attempts} attempts. Excel write skipped.")
    except Exception as e:
        print(f"Error preparing to write Excel: {e}")
        return False
    except Exception as e:
        print(f"Error writing to Excel: {e}")
        return False

    # Update CSV
    try:
        with open(CSV_PATH, "w", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=["waybill_no", "row", "timestamp"])
            writer.writeheader()
            writer.writerows(waybills)
    except Exception as e:
        print(f"Error writing to {CSV_PATH}: {e}")
        return False

    return True


if __name__ == "__main__":
    print("Generating waybills JSON and CSV...")
    if save_waybills_to_files():
        next_wb = get_next_unclaimed_waybill()
        if next_wb:
            print(f"Next unclaimed waybill: {next_wb}")
        else:
            print("All waybills have been claimed.")
    else:
        sys.exit(1)
