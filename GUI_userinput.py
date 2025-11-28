"""Two-trip waybill GUI with auto-load, reference lookup, and dual-sheet export to PDF.

Features:
- Auto-loads next unclaimed waybill from Database/waybills.json
- Collects 1st trip data (with 4 Reference no. and 2 Seal no. entries)
- Collects 2nd trip data (same structure)
- Writes both trips to print_1st2ndtrip.xlsx and WAYBILL RECORD.xlsx
- Exports to PDF with print area A1:O52
- Plate lookup with reference_data.xlsx

Run: `python GUI_userinput.py`
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime
import sys
import io
from pathlib import Path
import os
import time
import threading

# Import waybill functions
sys.path.insert(0, str(Path(__file__).parent))
from Database.read_waybills import get_next_unclaimed_waybill, mark_waybill_used, save_waybills_to_files


class TwoTripApp(tk.Tk):
	def __init__(self):
		super().__init__()
		self.title("Two-Trip Waybill Entry")
		self.geometry("900x800")
		self.current_waybill = None
		self.trip1_data = None
		self.trip2_data = None
		self._reference_cache = None
		
		# Redirect stdout to capture print statements
		self.log_queue = []
		self.original_stdout = sys.stdout
		
		self.create_widgets()
		self.refresh_waybill_database()
		self.load_next_waybill()
		self.show_trip1_frame()

	def create_widgets(self):
		"""Create main container frame."""
		self.main_frame = ttk.Frame(self, padding=10)
		self.main_frame.grid(row=0, column=0, sticky="nsew")
		self.columnconfigure(0, weight=1)
		self.rowconfigure(0, weight=1)
		self.main_frame.columnconfigure(0, weight=1)
		self.main_frame.rowconfigure(0, weight=1)
		
		# Log frame at the bottom
		log_frm = ttk.LabelFrame(self, text="Log", padding=5)
		log_frm.grid(row=1, column=0, sticky="nsew", padx=10, pady=(5, 10))
		self.columnconfigure(1, weight=1)
		self.rowconfigure(1, weight=0)
		
		self.log_text = tk.Text(log_frm, height=8, width=100, state=tk.DISABLED)
		self.log_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
		
		log_scroll = ttk.Scrollbar(log_frm, orient=tk.VERTICAL, command=self.log_text.yview)
		log_scroll.pack(side=tk.RIGHT, fill=tk.Y)
		self.log_text.config(yscrollcommand=log_scroll.set)

	def show_trip1_frame(self):
		"""Show 1st trip input frame."""
		# Clear main frame
		for widget in self.main_frame.winfo_children():
			widget.destroy()

		frm = ttk.LabelFrame(self.main_frame, text="1st Trip", padding=10)
		frm.grid(row=0, column=0, sticky="nsew", padx=5, pady=5)

		self.trip1_fields = self.create_trip_input_frame(frm, trip_num=1)

		# Buttons
		btn_frm = ttk.Frame(frm)
		btn_frm.grid(row=100, column=0, columnspan=4, pady=10)
		ttk.Button(btn_frm, text="Confirm 1st Trip", command=self.confirm_trip1).pack(side=tk.LEFT, padx=5)
		ttk.Button(btn_frm, text="Clear", command=lambda: self.clear_trip_fields(self.trip1_fields)).pack(side=tk.LEFT, padx=5)

	def show_trip2_frame(self):
		"""Show 2nd trip input frame after 1st trip confirmed."""
		# Clear main frame
		for widget in self.main_frame.winfo_children():
			widget.destroy()

		# Show 1st trip summary
		summary1_frm = ttk.LabelFrame(self.main_frame, text="1st Trip (Confirmed)", padding=8)
		summary1_frm.grid(row=0, column=0, sticky="nsew", padx=5, pady=5)
		self.show_trip_summary(summary1_frm, self.trip1_data, trip_num=1)

		# Show 2nd trip input
		frm = ttk.LabelFrame(self.main_frame, text="2nd Trip", padding=10)
		frm.grid(row=0, column=1, sticky="nsew", padx=5, pady=5)

		self.trip2_fields = self.create_trip_input_frame(frm, trip_num=2)

		# Buttons
		btn_frm = ttk.Frame(frm)
		btn_frm.grid(row=100, column=0, columnspan=4, pady=10)
		ttk.Button(btn_frm, text="Confirm 2nd Trip", command=self.confirm_trip2).pack(side=tk.LEFT, padx=5)
		ttk.Button(btn_frm, text="Clear", command=lambda: self.clear_trip_fields(self.trip2_fields)).pack(side=tk.LEFT, padx=5)
		ttk.Button(btn_frm, text="Back to 1st", command=self.show_trip1_frame).pack(side=tk.LEFT, padx=5)

	def show_final_summary(self):
		"""Show both trips confirmed with Print/Save/Clear options."""
		# Clear main frame
		for widget in self.main_frame.winfo_children():
			widget.destroy()

		title_lbl = ttk.Label(self.main_frame, text="Both Trips Confirmed", font=("Arial", 14, "bold"))
		title_lbl.grid(row=0, column=0, columnspan=2, sticky="w", padx=5, pady=5)

		# 1st trip
		summary1_frm = ttk.LabelFrame(self.main_frame, text="1st Trip", padding=8)
		summary1_frm.grid(row=1, column=0, sticky="nsew", padx=5, pady=5)
		self.show_trip_summary(summary1_frm, self.trip1_data, trip_num=1)

		# 2nd trip
		summary2_frm = ttk.LabelFrame(self.main_frame, text="2nd Trip", padding=8)
		summary2_frm.grid(row=1, column=1, sticky="nsew", padx=5, pady=5)
		self.show_trip_summary(summary2_frm, self.trip2_data, trip_num=2)

		# Action buttons
		btn_frm = ttk.Frame(self.main_frame)
		btn_frm.grid(row=2, column=0, columnspan=2, pady=15)
		ttk.Button(btn_frm, text="Export to PDF", command=self.export_to_pdf).pack(side=tk.LEFT, padx=5)
		ttk.Button(btn_frm, text="Clear All", command=self.clear_all_and_restart).pack(side=tk.LEFT, padx=5)

	def create_trip_input_frame(self, parent, trip_num):
		"""Create input frame for a trip (1 or 2)."""
		fields = {}

		# Waybill
		ttk.Label(parent, text="Waybill no:").grid(row=0, column=0, sticky="w", pady=2)
		ent_wb = ttk.Entry(parent, width=20)
		ent_wb.grid(row=0, column=1, sticky="w", pady=2)
		fields["waybill_no"] = ent_wb

		# Trip ticket (under waybill)
		ttk.Label(parent, text="Trip ticket no:").grid(row=1, column=0, sticky="w", pady=2)
		ent_tt = ttk.Entry(parent, width=20)
		ent_tt.grid(row=1, column=1, sticky="w", pady=2)
		fields["trip_ticket"] = ent_tt

		# Date
		ttk.Label(parent, text="Date:").grid(row=2, column=0, sticky="w", pady=2)
		ent_date = ttk.Entry(parent, width=20)
		ent_date.grid(row=2, column=1, sticky="w", pady=2)
		fields["date"] = ent_date

		# Plate no with Lookup button
		ttk.Label(parent, text="Plate no:").grid(row=3, column=0, sticky="w", pady=2)
		plate_frm = ttk.Frame(parent)
		plate_frm.grid(row=3, column=1, sticky="w", pady=2)
		ent_plate = ttk.Entry(plate_frm, width=15)
		ent_plate.pack(side=tk.LEFT, padx=(0, 4))
		ttk.Button(plate_frm, text="Lookup", width=8, command=lambda: self.lookup_plate(fields)).pack(side=tk.LEFT)
		fields["plate_no"] = ent_plate

		# Origin
		ttk.Label(parent, text="Origin:").grid(row=4, column=0, sticky="w", pady=2)
		ent_origin = ttk.Entry(parent, width=20)
		ent_origin.grid(row=4, column=1, sticky="w", pady=2)
		fields["origin"] = ent_origin

		# Reference no (4 entries)
		ttk.Label(parent, text="Reference no (up to 4):").grid(row=5, column=0, sticky="nw", pady=2)
		ref_frm = ttk.Frame(parent)
		ref_frm.grid(row=5, column=1, sticky="w", pady=2)
		ref_entries = []
		for i in range(4):
			ent = ttk.Entry(ref_frm, width=10)
			ent.pack(side=tk.LEFT, padx=2)
			ref_entries.append(ent)
		fields["reference_nos"] = ref_entries

		# Seal no (2 entries)
		ttk.Label(parent, text="Seal no (up to 2):").grid(row=6, column=0, sticky="nw", pady=2)
		seal_frm = ttk.Frame(parent)
		seal_frm.grid(row=6, column=1, sticky="w", pady=2)
		seal_entries = []
		for i in range(2):
			ent = ttk.Entry(seal_frm, width=10)
			ent.pack(side=tk.LEFT, padx=2)
			seal_entries.append(ent)
		fields["seal_nos"] = seal_entries

		# Total blocks
		ttk.Label(parent, text="Total Blcks:").grid(row=7, column=0, sticky="w", pady=2)
		ent_tb = ttk.Entry(parent, width=20)
		ent_tb.grid(row=7, column=1, sticky="w", pady=2)
		fields["total_blocks"] = ent_tb

		# Driver name
		ttk.Label(parent, text="Driver Name:").grid(row=8, column=0, sticky="w", pady=2)
		ent_driver = ttk.Entry(parent, width=20)
		ent_driver.grid(row=8, column=1, sticky="w", pady=2)
		fields["driver_name"] = ent_driver

		# Helper 1
		ttk.Label(parent, text="Helper 1:").grid(row=9, column=0, sticky="w", pady=2)
		ent_h1 = ttk.Entry(parent, width=20)
		ent_h1.grid(row=9, column=1, sticky="w", pady=2)
		fields["helper1"] = ent_h1

		# Helper 2
		ttk.Label(parent, text="Helper 2:").grid(row=10, column=0, sticky="w", pady=2)
		ent_h2 = ttk.Entry(parent, width=20)
		ent_h2.grid(row=10, column=1, sticky="w", pady=2)
		fields["helper2"] = ent_h2

		return fields

	def show_trip_summary(self, parent, trip_data, trip_num):
		"""Display a read-only summary of trip data."""
		if not trip_data:
			return
		summary_text = f"""Waybill: {trip_data.get('waybill_no', '')}
Trip Ticket: {trip_data.get('trip_ticket', '')}
Date: {trip_data.get('date', '')}
Plate: {trip_data.get('plate_no', '')}
Origin: {trip_data.get('origin', '')}
Ref: {trip_data.get('reference_no_formatted', '')}
Seal: {trip_data.get('seal_no_formatted', '')}
Blocks: {trip_data.get('total_blocks', '')}
Driver: {trip_data.get('driver_name', '')}
H1: {trip_data.get('helper1', '')}
H2: {trip_data.get('helper2', '')}"""
		text_widget = tk.Text(parent, width=35, height=12, state=tk.DISABLED)
		text_widget.grid(row=0, column=0, sticky="nsew")
		text_widget.config(state=tk.NORMAL)
		text_widget.insert("1.0", summary_text)
		text_widget.config(state=tk.DISABLED)

	def validate_trip_data(self, fields):
		"""Validate trip input; return dict or None if invalid."""
		data = {}
		for key, ent in fields.items():
			if key in ["reference_nos", "seal_nos"]:
				# Collect multi-entry fields
				values = [e.get().strip() for e in ent]
				values = [v for v in values if v]  # remove empties
				data[key] = values
			else:
				data[key] = ent.get().strip()

		# Validate required fields
		if not data.get("waybill_no"):
			messagebox.showwarning("Missing", "Please enter Waybill no.")
			return None
		if not data.get("trip_ticket"):
			messagebox.showwarning("Missing", "Please enter Trip ticket no.")
			return None
		if not data.get("date"):
			messagebox.showwarning("Missing", "Please enter Date.")
			return None
		try:
			datetime.strptime(data["date"], "%m-%d-%Y")
		except ValueError:
			messagebox.showwarning("Invalid", "Date must be mm-dd-yyyy.")
			return None
		if not data.get("plate_no"):
			messagebox.showwarning("Missing", "Please enter Plate no.")
			return None
		if not data.get("origin"):
			messagebox.showwarning("Missing", "Please enter Origin.")
			return None
		if not data.get("driver_name"):
			messagebox.showwarning("Missing", "Please enter Driver Name.")
			return None

		# Convert total blocks
		tb = data.get("total_blocks", "")
		try:
			data["total_blocks"] = int(tb) if tb else 0
		except ValueError:
			messagebox.showwarning("Invalid", "Total Blcks must be integer.")
			return None

		# Format reference no (max 3 shown, ellipsis if > 3)
		ref_list = data.get("reference_nos", [])
		if len(ref_list) > 3:
			data["reference_no_formatted"] = "/".join(ref_list[:3]) + "…"
		else:
			data["reference_no_formatted"] = "/".join(ref_list) if ref_list else ""

		# Format seal no (join up to 2)
		seal_list = data.get("seal_nos", [])
		data["seal_no_formatted"] = "/".join(seal_list[:2]) if seal_list else ""

		return data

	def confirm_trip1(self):
		"""Confirm 1st trip: validate, store, write to Excel, show 2nd trip input."""
		data = self.validate_trip_data(self.trip1_fields)
		if not data:
			return
		self.trip1_data = data
		self.log_print(f"Trip 1 confirmed: {data}")

		try:
			self.write_trip_to_excel(data, trip_num=1)
			# Write first trip data to WAYBILL RECORD for the first waybill
			self.write_to_waybill_record(data, waybill_no=data.get("waybill_no"))
		except Exception as e:
			messagebox.showerror("Error", f"Could not write to Excel: {e}")
			self.log_print(f"Error: {e}")
			return

		# Move to 2nd trip input frame and auto-fill 2nd waybill = 1st waybill + 1 when possible
		self.show_trip2_frame()
		# Auto-increment waybill for trip 2 if numeric
		try:
			wb1 = data.get("waybill_no", "")
			if wb1 and str(wb1).isdigit():
				next_wb = str(int(wb1) + 1)
				if hasattr(self, 'trip2_fields') and self.trip2_fields.get('waybill_no'):
					self.trip2_fields['waybill_no'].delete(0, tk.END)
					self.trip2_fields['waybill_no'].insert(0, next_wb)
					self.log_print(f"Auto-filled 2nd trip waybill: {next_wb}")
		except Exception:
			# Non-fatal; just skip auto-fill
			pass

	def confirm_trip2(self):
		"""Confirm 2nd trip: validate, store, then show final options."""
		data = self.validate_trip_data(self.trip2_fields)
		if not data:
			return
		self.trip2_data = data
		self.log_print(f"Trip 2 confirmed: {data}")

		try:
			self.write_trip_to_excel(data, trip_num=2)
			# Write second trip data to WAYBILL RECORD for the second waybill
			self.write_to_waybill_record(data, waybill_no=data.get("waybill_no"))
		except Exception as e:
			messagebox.showerror("Error", f"Could not write to Excel: {e}")
			self.log_print(f"Error: {e}")
			return

		# Mark both waybills as used now that both trips are done (if present)
		def _mark_async(w1, w2):
			try:
				if w1:
					res1 = mark_waybill_used(w1)
					self.log_print(f"Marked waybill {w1} as used: {res1}")
					# small delay between marks
					time.sleep(0.2)
				if w2 and w2 != w1:
					res2 = mark_waybill_used(w2)
					self.log_print(f"Marked waybill {w2} as used: {res2}")
			except Exception as e:
				self.log_print(f"Warning: Could not mark waybills as used: {e}")

		wb1 = self.trip1_data.get("waybill_no") if self.trip1_data else None
		wb2 = self.trip2_data.get("waybill_no") if self.trip2_data else None
		threading.Thread(target=_mark_async, args=(wb1, wb2), daemon=True).start()

		self.show_final_summary()

	def write_trip_to_excel(self, data, trip_num):
		"""Write trip data to print_1st2ndtrip.xlsx at specific cells."""
		try:
			import openpyxl
		except Exception:
			raise Exception("openpyxl required: python -m pip install openpyxl")

		ref_path = Path(__file__).parent / "reference" / "print_1st2ndtrip.xlsx"
		if not ref_path.exists():
			raise Exception(f"Print template not found: {ref_path}")

		wb = openpyxl.load_workbook(ref_path)
		ws = wb.active

		if trip_num == 1:
			# 1st trip cells
			ws["N2"] = data.get("waybill_no", "")
			ws["M4"] = data.get("date", "")
			ws["G5"] = data.get("plate_no", "")
			ws["P1"] = data.get("origin", "")
			ws["J11"] = data.get("trip_ticket", "")
			ws["D11"] = data.get("reference_no_formatted", "")
			# Seal no: J12 and K12 (moved 1 cell right)
			seals = data.get("seal_nos", [])
			if len(seals) > 0:
				ws["J12"] = seals[0]
			if len(seals) > 1:
				ws["K12"] = seals[1]
			ws["C17"] = data.get("total_blocks", 0)
			ws["C5"] = data.get("driver_name", "")
			ws["J5"] = data.get("helper1", "")
			ws["K5"] = data.get("helper2", "")
		else:  # trip_num == 2
			# 2nd trip cells
			ws["N28"] = data.get("waybill_no", "")
			ws["M30"] = data.get("date", "")
			ws["G31"] = data.get("plate_no", "")
			ws["P27"] = data.get("origin", "")
			ws["J37"] = data.get("trip_ticket", "")
			ws["D37"] = data.get("reference_no_formatted", "")
			# Seal no: J38 and K38 (moved 1 cell right)
			seals = data.get("seal_nos", [])
			if len(seals) > 0:
				ws["J38"] = seals[0]
			if len(seals) > 1:
				ws["K38"] = seals[1]
			ws["C43"] = data.get("total_blocks", 0)
			ws["C31"] = data.get("driver_name", "")
			ws["J31"] = data.get("helper1", "")
			ws["K31"] = data.get("helper2", "")

		wb.save(ref_path)
		self.log_print(f"Wrote trip {trip_num} to {ref_path}")

	def write_to_waybill_record(self, data, waybill_no=None):
		"""Write trip data to WAYBILL RECORD.xlsx in row matching given waybill_no.
		If waybill_no is None, uses self.current_waybill.
		"""
		if waybill_no is None:
			waybill_no = self.current_waybill
		try:
			import openpyxl
		except Exception:
			raise Exception("openpyxl required: python -m pip install openpyxl")

		waybill_path = Path(__file__).parent / "Database" / "WAYBILL RECORD.xlsx"
		if not waybill_path.exists():
			raise Exception(f"WAYBILL RECORD not found: {waybill_path}")

		wb = openpyxl.load_workbook(waybill_path)
		ws = wb.active

		# Find row matching waybill_no in column A
		wb_row = None
		for r in range(1, ws.max_row + 1):
			cell_val = ws.cell(row=r, column=1).value
			if cell_val and str(cell_val).strip() == str(waybill_no):
				wb_row = r
				break

		if not wb_row:
			self.log_print(f"Warning: Could not find waybill {waybill_no} in WAYBILL RECORD")
			return

		# Write to specific columns
		# Column C per user request should indicate TRANSFER instead of timestamp
		ws.cell(row=wb_row, column=3).value = "TRANSFER"  # C
		ws.cell(row=wb_row, column=4).value = data.get("total_blocks", 0)  # D
		ws.cell(row=wb_row, column=5).value = data.get("trip_ticket", "")  # E
		ws.cell(row=wb_row, column=7).value = data.get("date", "")  # G
		ws.cell(row=wb_row, column=9).value = data.get("driver_name", "")  # I
		ws.cell(row=wb_row, column=10).value = data.get("helper1", "")  # J
		ws.cell(row=wb_row, column=11).value = data.get("helper2", "")  # K
		ws.cell(row=wb_row, column=12).value = data.get("origin", "")  # L

		# Attempt to save workbook with retries to handle brief locks (like Excel open)
		max_attempts = 5
		for attempt in range(1, max_attempts + 1):
			try:
				wb.save(waybill_path)
				self.log_print(f"Wrote to WAYBILL RECORD row {wb_row}")
				break
			except Exception as e:
				self.log_print(f"Attempt {attempt}/{max_attempts} failed to save WAYBILL RECORD: {e}")
				if attempt < max_attempts:
					time.sleep(1)
				else:
					self.log_print("Failed to write WAYBILL RECORD after several attempts. Please close the Excel file and try again.")

	def export_to_pdf(self):
		"""Open the filled print Excel and then clear the template for next use."""
		try:
			ref_path = Path(__file__).parent / "reference" / "print_1st2ndtrip.xlsx"
			wb1 = self.trip1_data.get("waybill_no", "") if self.trip1_data else ""
			wb2 = self.trip2_data.get("waybill_no", "") if self.trip2_data else ""

			if not ref_path.exists():
				messagebox.showerror("Error", f"Print template/file not found: {ref_path}")
				return

			# Open the Excel file for the user (Windows)
			try:
				os.startfile(str(ref_path))
				self.log_print(f"Opened {ref_path} in associated application.")
			except Exception as e:
				self.log_print(f"Could not open file automatically: {e}")

			# Clear the GUI (not the Excel file) so the form is ready for next entry
			self.clear_all_and_restart()

			messagebox.showinfo("Done", f"Opened {ref_path}. GUI cleared for next entry.")
		except Exception as e:
			messagebox.showerror("Error", f"Could not finalize export: {e}")
			self.log_print(f"Error: {e}")

	def clear_print_template(self):
		"""Clear the cells in the print template used for trips so it's ready for next use."""
		try:
			import openpyxl
		except Exception:
			self.log_print("openpyxl required to clear template.")
			return

		ref_path = Path(__file__).parent / "reference" / "print_1st2ndtrip.xlsx"
		if not ref_path.exists():
			self.log_print(f"Template not found to clear: {ref_path}")
			return

		wb = openpyxl.load_workbook(ref_path)
		ws = wb.active
		# Clear cells written by write_trip_to_excel
		cells_to_clear = [
			# Trip 1
			"N2","M4","G5","P1","J11","D11","J12","K12","C17","C5","J5","K5",
			# Trip 2
			"N28","M30","G31","P27","J37","D37","J38","K38","C43","C31","J31","K31",
		]
		for c in cells_to_clear:
			try:
				ws[c] = None
			except Exception:
				pass
		wb.save(ref_path)
		self.log_print(f"Cleared template {ref_path}")

	def export_to_pdf(self):
		"""Overwrite print_1st2ndtrip.xlsx with the trip data (already written by write_trip_to_excel)."""
		try:
			ref_path = Path(__file__).parent / "reference" / "print_1st2ndtrip.xlsx"
			wb1 = self.trip1_data.get("waybill_no", "")
			wb2 = self.trip2_data.get("waybill_no", "")
			
			# The data is already written to the Excel file by write_trip_to_excel calls
			# Just confirm to the user that the file is ready
			self.log_print(f"Data written to {ref_path}")
			self.log_print(f"Ready for manual conversion to PDF: {wb1}-{wb2}.pdf")
			messagebox.showinfo("Success", f"Data saved to:\n{ref_path}\n\nReady to convert to {wb1}-{wb2}.pdf manually or via LibreOffice.")
		except Exception as e:
			messagebox.showerror("Error", f"Could not finalize export: {e}")
			self.log_print(f"Error: {e}")

	def clear_all_and_restart(self):
		"""Clear both trips and restart."""
		self.trip1_data = None
		self.trip2_data = None
		self.load_next_waybill()
		self.show_trip1_frame()

	def clear_trip_fields(self, fields):
		"""Clear all input fields in a trip."""
		for key, ent in fields.items():
			if isinstance(ent, list):
				for e in ent:
					e.delete(0, tk.END)
			else:
				ent.delete(0, tk.END)

	def load_next_waybill(self):
		"""Load next unclaimed waybill."""
		try:
			next_wb = get_next_unclaimed_waybill()
			if next_wb:
				self.current_waybill = next_wb
				self.log_print(f"Loaded waybill: {next_wb}")
			else:
				messagebox.showinfo("Info", "All waybills claimed.")
				self.log_print("All waybills claimed.")
		except Exception as e:
			messagebox.showerror("Error", f"Could not load waybill: {e}")
			self.log_print(f"Error loading waybill: {e}")

	def refresh_waybill_database(self):
		"""Refresh waybill database by running read_waybills.py."""
		try:
			self.log_print("Refreshing waybill database...")
			save_waybills_to_files()
			self.log_print("Waybill database refreshed.")
		except Exception as e:
			self.log_print(f"Warning: Could not refresh database: {e}")

	def log_print(self, msg):
		"""Print to both terminal and GUI log box."""
		print(msg)  # Also print to terminal
		self.log_text.config(state=tk.NORMAL)
		self.log_text.insert(tk.END, msg + "\n")
		self.log_text.see(tk.END)
		self.log_text.config(state=tk.DISABLED)
		self.update_idletasks()

	def normalize_plate(self, plate: str) -> str:
		"""Normalize plate: remove spaces, uppercase."""
		if not plate:
			return ""
		return plate.replace(" ", "").upper()

	def load_reference_data(self):
		"""Load reference_data.xlsx (cached)."""
		if self._reference_cache is not None:
			return self._reference_cache
		ref_path = Path(__file__).parent / "Database" / "reference_data.xlsx"
		if not ref_path.exists():
			self.log_print(f"Reference file not found: {ref_path}")
			self._reference_cache = []
			return self._reference_cache
		try:
			import openpyxl
		except Exception:
			messagebox.showerror("Missing", "openpyxl required: python -m pip install openpyxl")
			self._reference_cache = []
			return self._reference_cache

		wb = openpyxl.load_workbook(ref_path, data_only=True)
		ws = wb.active
		items = []
		for r in range(1, ws.max_row + 1):
			truck = ws.cell(row=r, column=1).value
			if not truck:
				continue
			driver = ws.cell(row=r, column=2).value
			helper1 = ws.cell(row=r, column=3).value
			helper2 = ws.cell(row=r, column=4).value
			item = {
				"truck_raw": str(truck).strip(),
				"truck_norm": self.normalize_plate(str(truck)),
				"driver": (str(driver).strip() if driver else ""),
				"helper1": (str(helper1).strip() if helper1 else ""),
				"helper2": (str(helper2).strip() if helper2 else ""),
			}
			items.append(item)
		self._reference_cache = items
		self.log_print(f"Loaded {len(items)} reference rows")
		return items

	def show_selection_dialog(self, matches):
		"""Show selection dialog for multiple matches."""
		dlg = tk.Toplevel(self)
		dlg.title("Select match")
		dlg.geometry("420x200")
		lb = tk.Listbox(dlg, width=80, height=8)
		for m in matches:
			text = f"{m['truck_raw']}  —  {m['driver']}  |  {m['helper1']}  |  {m['helper2']}"
			lb.insert(tk.END, text)
		lb.pack(padx=8, pady=8, fill=tk.BOTH, expand=True)

		chosen = {"item": None}

		def on_ok():
			sel = lb.curselection()
			if not sel:
				return
			chosen['item'] = matches[sel[0]]
			dlg.destroy()

		ttk.Button(dlg, text="OK", command=on_ok).pack(pady=(0, 8))
		dlg.transient(self)
		dlg.grab_set()
		self.wait_window(dlg)
		return chosen['item']

	def lookup_plate(self, fields):
		"""Lookup plate in reference data."""
		plate = fields.get("plate_no").get().strip()
		if not plate:
			self.log_print("Lookup: Plate is empty")
			return
		norm = self.normalize_plate(plate)
		refs = self.load_reference_data()
		if not refs:
			self.log_print("Lookup: No reference data")
			return
		matches = [r for r in refs if r.get("truck_norm") == norm]
		if not matches:
			self.log_print(f"Lookup: No match for '{plate}'")
			return
		if len(matches) == 1:
			m = matches[0]
			self.log_print(f"Lookup: {m['truck_raw']} -> {m['driver']}, {m['helper1']}, {m['helper2']}")
			fields['driver_name'].delete(0, tk.END)
			fields['driver_name'].insert(0, m.get('driver', ''))
			fields['helper1'].delete(0, tk.END)
			fields['helper1'].insert(0, m.get('helper1', ''))
			fields['helper2'].delete(0, tk.END)
			fields['helper2'].insert(0, m.get('helper2', ''))
			return
		choice = self.show_selection_dialog(matches)
		if choice:
			self.log_print(f"Lookup: {choice['truck_raw']} -> {choice['driver']}, {choice['helper1']}, {choice['helper2']}")
			fields['driver_name'].delete(0, tk.END)
			fields['driver_name'].insert(0, choice.get('driver', ''))
			fields['helper1'].delete(0, tk.END)
			fields['helper1'].insert(0, choice.get('helper1', ''))
			fields['helper2'].delete(0, tk.END)
			fields['helper2'].insert(0, choice.get('helper2', ''))
		else:
			self.log_print("Lookup: No selection")


if __name__ == "__main__":
	app = TwoTripApp()
	app.mainloop()

