"""Inspect reference_data.xlsx structure."""
import openpyxl
from pathlib import Path

ref_file = Path(__file__).parent / "Database" / "reference_data.xlsx"
wb = openpyxl.load_workbook(ref_file, data_only=True)
ws = wb.active

print("Columns:", [ws.cell(1, i).value for i in range(1, ws.max_column + 1)])
print("\nFirst 10 rows:")
for r in range(1, min(11, ws.max_row + 1)):
    row_data = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
    print(f"Row {r}: {row_data}")
