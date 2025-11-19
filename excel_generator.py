# excel_generator.py
# FINAL WEAPON — ONE CLICK → PERFECT LTS FILE
# Lead Engineer: Marlou V. Bation
# Day 4 — 20 Nov 2025

from pathlib import Path
from datetime import datetime
import openpyxl
from openpyxl import load_workbook

TEMPLATE = Path("templates/print_1st2ndtrip.xlsx")
OUTPUT_DIR = Path("output")
OUTPUT_DIR.mkdir(exist_ok=True)

def generate_lts_excel(data: dict):
    # Auto-increment LTS number: LTS097, LTS098...
    existing = sorted(OUTPUT_DIR.glob("*.xlsx"))
    next_num = 97 if not existing else (97 + len(existing))
    today = datetime.now().strftime("%d-%b-%Y").upper()
    filename = OUTPUT_DIR / f"{today} (LTS{chr(next_num)}).xlsx"

    wb = load_workbook(TEMPLATE, keep_vba=False, data_only=True)
    ws = wb.active

    # === EXACT CELL MAPPING FROM YOUR TEMPLATE (PAGE 1 / FIRST TRIP) ===
    mapping = {
        "trip_ticket":     "J11",   # Trip Ticket No.
        "delivery_date":   "M4",    # Date
        "plate_no":        "G5",    # Plate No.
        "driver":          "C5",    # Driver name (you had G56 — that was wrong, it's C5)
        "helper1":         "I5",    # Helper name
        "total_blocks":    "C17",   # RM TRANSFER blocks
        "ref_nos":         "D11",   # Ref Nos (V5176 etc.)
        "seal_nos":        "J12",   # Seal No.
        "shipper_full":    "C6",    # Shipper (VIRGINIA FOOD, INC.)
        "from_location":   "J6",    # From:
        "to_location":     "J7",    # To:
    }

    for key, cell in mapping.items():
        value = data.get(key, "")
        if value and value != "[empty]":
            ws[cell] = value

    # Optional: Write date in bottom signature line too
    if data.get("delivery_date"):
        ws["G23"] = data["delivery_date"]  # Date under signature

    wb.save(filename)
    print(f"SUCCESS → {filename.name}")
    return filename